"""Shared visual language for the FounderOS workbook.

Central palette + helper builders so every sheet module renders with the
same "premium software" look instead of ad-hoc formatting per sheet.
"""
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Segoe UI"

# ---- Palette -----------------------------------------------------------
NAVY = "1F2937"
NAVY_DARK = "111827"
SLATE = "334155"
GRAY = "6B7280"
BORDER_GRAY = "D9DEE4"
BG_LIGHT = "F8FAFC"
WHITE = "FFFFFF"

ACCENT_BLUE = "2563EB"
ACCENT_TEAL = "0D9488"
ACCENT_PURPLE = "7C3AED"
ACCENT_AMBER = "D97706"
ACCENT_ROSE = "DB2777"

SUCCESS = "16A34A"
WARNING = "F59E0B"
DANGER = "DC2626"
INFO = "2563EB"

MODULE_COLORS = {
    "home": NAVY,
    "companies": ACCENT_TEAL,
    "projects": ACCENT_PURPLE,
    "tasks": ACCENT_AMBER,
    "time": ACCENT_ROSE,
    "employees": ACCENT_BLUE,
    "lookups": GRAY,
}


def hex_fill(hex_color, fg=True):
    return PatternFill("solid", fgColor=hex_color) if fg else PatternFill("solid", bgColor=hex_color)


def header_font(color=WHITE, size=11, bold=True):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def body_font(color=NAVY_DARK, size=10, bold=False, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=color)


THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)


def style_table_header_row(ws, row, first_col, last_col, module_key):
    fill = hex_fill(MODULE_COLORS.get(module_key, NAVY))
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font()
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def add_table(ws, ref, name, style="TableStyleMedium9"):
    """Register an Excel structured Table over `ref` (e.g. 'A1:H10')."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    return tbl


def autosize_columns(ws, widths):
    """widths: dict col_letter -> width, applied directly (fast, deterministic)."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def apply_status_conditional_formatting(ws, cell_range, mapping):
    """mapping: {text: hex_color} -> colors whole cell when it equals text."""
    for text, color in mapping.items():
        rule = FormulaRule(
            formula=[f'EXACT({cell_range.split(":")[0]},"{text}")'],
            fill=hex_fill(color),
            font=Font(name=FONT_NAME, color=WHITE, bold=True),
        )
        ws.conditional_formatting.add(cell_range, rule)


def add_databar(ws, cell_range, color=ACCENT_BLUE):
    rule = DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=100,
        color=color, showValue=True, minLength=None, maxLength=None,
    )
    ws.conditional_formatting.add(cell_range, rule)


def add_data_validation_list(ws, cell_range, formula1, allow_blank=True):
    """formula1 like '=List_Departments' (a defined name) or '=Sheet!$A$1:$A$5'."""
    dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank, showDropDown=False)
    dv.error = "Please choose a value from the dropdown list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def set_col_widths_auto(ws, min_row, max_row, first_col, last_col, pad=2, cap=42, floor=10):
    for c in range(first_col, last_col + 1):
        letter = get_column_letter(c)
        longest = len(str(ws.cell(row=min_row, column=c).value or ""))
        for r in range(min_row + 1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = max(floor, min(cap, longest + pad))
