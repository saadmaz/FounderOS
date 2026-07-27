"""Entry point: generates FounderOS.xlsx from scratch.

Run with:  python -m founderos.build
Regenerate any time the sheet modules or seed data change — the whole
workbook is rebuilt deterministically, nothing is hand-edited in place.
"""
import os
from openpyxl import Workbook

from .sheets import (
    lookups, vendors, companies, employees, projects, tasks, time_log,
    expenses, revenue, invoices, budgets, finance, home,
)

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output", "FounderOS.xlsx")


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    # Build order matters only for defined-name / cross-sheet-object
    # dependencies at *generation* time; on-screen tab order is set
    # explicitly via each create_sheet(..., index=...) call.
    lookups.build(wb)
    vendors.build(wb)
    companies.build(wb)
    employees.build(wb)
    projects.build(wb)
    tasks.build(wb)
    time_log.build(wb)
    expenses.build(wb)
    revenue.build(wb)
    invoices.build(wb)
    budgets.build(wb)
    finance.build(wb)
    home.build(wb)  # index=0 -> pinned as the first (leftmost) tab

    wb.active = wb.sheetnames.index("Home")
    wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build_workbook()
    print(f"FounderOS workbook written to: {path}")
