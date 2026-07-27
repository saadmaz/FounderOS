"""Expenses: the transactional master table finance KPIs (P&L, runway,
outstanding reimbursements) roll up from. Project is optional - most
overhead expenses aren't tied to a single project.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import COMPANIES, PROJECTS, EXPENSES
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Expenses"

COLUMNS = [
    ("Expense ID", 12), ("Company", 18), ("Project", 22), ("Category", 18),
    ("Vendor", 20), ("Date", 12), ("Amount", 11), ("Currency", 10), ("Tax", 9),
    ("Total", 11), ("Payment Method", 14), ("Paid By", 18), ("Reimbursable", 12),
    ("Reimbursed", 11), ("Reimbursement Date", 15), ("Approval Status", 13),
    ("Payment Status", 13), ("Receipt Link", 26), ("Receipt", 14), ("Notes", 26),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Expenses")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["tasks"]
    common.write_title(ws, "🧾  Expenses",
                        "One row per expense. Paste a receipt link (Drive/OneDrive/Dropbox/PDF URL) and the Receipt column becomes a one-click hyperlink.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "tasks")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    company_name = {f"COMP-{i+1:04d}": c["name"] for i, c in enumerate(COMPANIES)}
    project_name_by_index = {i + 1: p[1] for i, p in enumerate(PROJECTS)}

    for i, row in enumerate(EXPENSES):
        (comp_id, proj_idx, category, vendor, edate, amount, currency, tax,
         pay_method, paid_by, reimbursable, reimbursed, reimb_date, approval,
         pay_status, receipt_link, notes) = row
        r = r0 + i
        anchor_col = col_letter(idx["Date"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Expense ID"])
        id_cell.value = f'="EXP-"&TEXT(SUBTOTAL(103,${anchor_col}${r0}:{anchor_col}{r}),"0000")'

        vals = {
            "Company": company_name[comp_id],
            "Project": project_name_by_index.get(proj_idx, ""),
            "Category": category, "Vendor": vendor, "Date": edate, "Amount": amount,
            "Currency": currency, "Tax": tax, "Payment Method": pay_method, "Paid By": paid_by,
            "Reimbursable": reimbursable, "Reimbursed": reimbursed,
            "Reimbursement Date": reimb_date, "Approval Status": approval,
            "Payment Status": pay_status, "Receipt Link": receipt_link, "Notes": notes,
        }
        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field in ("Date", "Reimbursement Date"):
                cell.number_format = "yyyy-mm-dd"
            if field in ("Amount", "Tax"):
                cell.number_format = "#,##0.00"

        amt_ref = f"{col_letter(idx['Amount'])}{r}"
        tax_ref = f"{col_letter(idx['Tax'])}{r}"
        total_cell = ws.cell(row=r, column=FIRST_COL + idx["Total"], value=f'=ROUND({amt_ref}+{tax_ref},2)')
        total_cell.number_format = "#,##0.00"

        link_ref = f"{col_letter(idx['Receipt Link'])}{r}"
        ws.cell(row=r, column=FIRST_COL + idx["Receipt"],
                value=f'=IF({link_ref}="","",HYPERLINK({link_ref},"📎 Open Receipt"))')

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(EXPENSES) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Company": "=List_CompanyNames",
        "Project": "=List_ProjectNames",
        "Category": f"={NAME_MAP['Expense Categories']}",
        "Vendor": "=List_VendorNames",
        "Currency": f"={NAME_MAP['Currencies']}",
        "Payment Method": f"={NAME_MAP['Payment Methods']}",
        "Reimbursable": f"={NAME_MAP['Billable']}",
        "Reimbursed": f"={NAME_MAP['Billable']}",
        "Approval Status": f"={NAME_MAP['Approval Status']}",
        "Payment Status": f"={NAME_MAP['Payment Status']}",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}20000", formula)

    approval_col = col_letter(idx["Approval Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{approval_col}{r0}:{approval_col}20000",
        {"Pending": styles.WARNING, "Approved": styles.SUCCESS, "Rejected": styles.DANGER},
    )
    pay_col = col_letter(idx["Payment Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{pay_col}{r0}:{pay_col}20000",
        {"Unpaid": styles.DANGER, "Paid": styles.SUCCESS, "Partially Paid": styles.WARNING, "Overdue": styles.DANGER},
    )

    # Outstanding reimbursement highlight: Reimbursable=Yes and Reimbursed=No
    reimb_col = col_letter(idx["Reimbursable"])
    reimbd_col = col_letter(idx["Reimbursed"])
    from openpyxl.formatting.rule import FormulaRule
    rule = FormulaRule(
        formula=[f'AND(${reimb_col}{r0}="Yes",${reimbd_col}{r0}="No")'],
        fill=styles.hex_fill(styles.WARNING),
        font=styles.header_font(),
    )
    ws.conditional_formatting.add(f"{reimbd_col}{r0}:{reimbd_col}20000", rule)

    define_dynamic_range(wb, "List_ExpenseIDs", "Expenses", col_letter(idx["Expense ID"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Expense ID'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
