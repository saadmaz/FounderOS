"""Employees: master people table. Feeds the Owner/Manager/Employee
dropdowns on Projects, Tasks, and Time Log.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import EMPLOYEES
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Employees"

COLUMNS = [
    ("Employee ID", 12), ("Name", 20), ("Email", 26), ("Role / Title", 20),
    ("Department", 16), ("Primary Company", 20), ("Status", 11), ("Start Date", 12),
    ("Open Tasks", 11), ("Hours Logged", 12), ("Skills", 26), ("Notes", 26),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Employees")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["employees"]
    common.write_title(ws, "🧑‍💼  Employees — People Directory",
                        "One row per person across all companies. Referenced by Projects, Tasks, and Time Log as Owner/Employee.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "employees")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    # EMPLOYEES tuples: (name, email, role, department, primary_company_id, status, start_date, skills)
    # primary_company_id in seed_data is a "COMP-000x" placeholder; resolve to company name via COMPANIES order.
    from ..seed_data import COMPANIES
    comp_id_to_name = {f"COMP-{i+1:04d}": c["name"] for i, c in enumerate(COMPANIES)}

    for i, (name, email, role, dept, comp_id, status, start, skills) in enumerate(EMPLOYEES):
        r = r0 + i
        name_col = col_letter(idx["Name"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Employee ID"])
        id_cell.value = f'="EMP-"&TEXT(SUBTOTAL(103,${name_col}${r0}:{name_col}{r}),"0000")'

        vals = {
            "Name": name, "Email": email, "Role / Title": role, "Department": dept,
            "Primary Company": comp_id_to_name.get(comp_id, comp_id), "Status": status,
            "Start Date": start, "Skills": skills, "Notes": "",
        }
        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field == "Start Date":
                cell.number_format = "yyyy-mm-dd"

        emp_name_ref = f"{name_col}{r}"
        ws.cell(row=r, column=FIRST_COL + idx["Open Tasks"],
                value=f'=COUNTIFS(Tbl_Tasks[Owner],{emp_name_ref},Tbl_Tasks[Status],"<>Completed",Tbl_Tasks[Status],"<>Cancelled")')
        hrs = ws.cell(row=r, column=FIRST_COL + idx["Hours Logged"],
                       value=f'=ROUND(SUMIFS(Tbl_TimeLog[Total Hours],Tbl_TimeLog[Employee],{emp_name_ref}),1)')
        hrs.number_format = "0.0"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(EMPLOYEES) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    styles.add_data_validation_list(ws, f"{col_letter(idx['Department'])}{r0}:{col_letter(idx['Department'])}1000",
                                     f"={NAME_MAP['Departments']}")
    styles.add_data_validation_list(ws, f"{col_letter(idx['Primary Company'])}{r0}:{col_letter(idx['Primary Company'])}1000",
                                     "=List_CompanyNames")
    styles.add_data_validation_list(ws, f"{col_letter(idx['Status'])}{r0}:{col_letter(idx['Status'])}1000",
                                     f"={NAME_MAP['Employee Status']}")

    status_col = col_letter(idx["Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{status_col}{r0}:{status_col}1000",
        {"Active": styles.SUCCESS, "Inactive": styles.GRAY, "On Leave": styles.WARNING},
    )

    define_dynamic_range(wb, "List_EmployeeNames", "Employees", col_letter(idx["Name"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Name'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
