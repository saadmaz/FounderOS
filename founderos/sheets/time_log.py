"""Time Log: work sessions. Task Name / Project / Company are derived
automatically from the Task ID via INDEX/MATCH so logging time only
requires picking a Task and an Employee.
"""
from datetime import datetime
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import TASKS, TIME_LOG
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_TimeLog"

COLUMNS = [
    ("Log ID", 12), ("Task", 12), ("Task Name", 30), ("Project", 22), ("Company", 18),
    ("Employee", 18), ("Date", 12), ("Start Time", 11), ("End Time", 11),
    ("Break (min)", 11), ("Total Hours", 12), ("Billable", 10), ("Category", 14), ("Notes", 26),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Time_Log")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["time"]
    common.write_title(ws, "⏱️  Time Log",
                        "One row per work session. Task Name/Project/Company auto-fill from the selected Task.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "time")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    task_id_by_index = {i + 1: f"TASK-{i + 1:04d}" for i in range(len(TASKS))}

    for i, (task_idx, employee, log_date, start, end, brk, billable, category, notes) in enumerate(TIME_LOG):
        r = r0 + i
        date_col = col_letter(idx["Date"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Log ID"])
        id_cell.value = f'="TL-"&TEXT(SUBTOTAL(103,${date_col}${r0}:{date_col}{r}),"0000")'

        ws.cell(row=r, column=FIRST_COL + idx["Task"], value=task_id_by_index[task_idx])
        d_cell = ws.cell(row=r, column=FIRST_COL + idx["Date"], value=log_date)
        d_cell.number_format = "yyyy-mm-dd"
        st_cell = ws.cell(row=r, column=FIRST_COL + idx["Start Time"],
                           value=datetime.strptime(start, "%H:%M").time())
        st_cell.number_format = "hh:mm"
        et_cell = ws.cell(row=r, column=FIRST_COL + idx["End Time"],
                           value=datetime.strptime(end, "%H:%M").time())
        et_cell.number_format = "hh:mm"
        ws.cell(row=r, column=FIRST_COL + idx["Break (min)"], value=brk)
        ws.cell(row=r, column=FIRST_COL + idx["Employee"], value=employee)
        ws.cell(row=r, column=FIRST_COL + idx["Billable"], value=billable)
        ws.cell(row=r, column=FIRST_COL + idx["Category"], value=category)
        ws.cell(row=r, column=FIRST_COL + idx["Notes"], value=notes)

        task_ref = f"{col_letter(idx['Task'])}{r}"
        tn_cell = ws.cell(row=r, column=FIRST_COL + idx["Task Name"],
                           value=f'=IFERROR(INDEX(Tbl_Tasks[Task Name],MATCH({task_ref},Tbl_Tasks[Task ID],0)),"")')
        proj_cell = ws.cell(row=r, column=FIRST_COL + idx["Project"],
                             value=f'=IFERROR(INDEX(Tbl_Tasks[Project],MATCH({task_ref},Tbl_Tasks[Task ID],0)),"")')
        comp_cell = ws.cell(row=r, column=FIRST_COL + idx["Company"],
                             value=f'=IFERROR(INDEX(Tbl_Tasks[Company],MATCH({task_ref},Tbl_Tasks[Task ID],0)),"")')

        start_ref = f"{col_letter(idx['Start Time'])}{r}"
        end_ref = f"{col_letter(idx['End Time'])}{r}"
        break_ref = f"{col_letter(idx['Break (min)'])}{r}"
        total_cell = ws.cell(row=r, column=FIRST_COL + idx["Total Hours"],
                              value=f'=ROUND(MAX(0,(({end_ref}-{start_ref})*24)-({break_ref}/60)),2)')
        total_cell.number_format = "0.00"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(TIME_LOG) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Task": "=List_TaskIDs",
        "Employee": "=List_EmployeeNames",
        "Billable": f"={NAME_MAP['Billable']}",
        "Category": f"={NAME_MAP['Time Categories']}",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}20000", formula)

    ws.freeze_panes = f"{col_letter(idx['Task'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
