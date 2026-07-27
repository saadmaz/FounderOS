"""Tasks: joins to Projects by name (Company is auto-derived via
INDEX/MATCH so nobody has to pick it twice). Time Log rolls up into
Actual Hours and Progress %.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import PROJECTS, TASKS, COMPANIES
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Tasks"

COLUMNS = [
    ("Task ID", 12), ("Task Name", 32), ("Project", 24), ("Company", 18), ("Department", 15),
    ("Owner", 18), ("Task Type", 12), ("Priority", 11), ("Status", 13),
    ("Created Date", 13), ("Due Date", 12), ("Completed Date", 14),
    ("Estimated Hours", 13), ("Actual Hours", 12), ("Progress %", 11), ("Notes", 28),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Tasks")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["tasks"]
    common.write_title(ws, "✅  Tasks",
                        "One row per task. Company is derived automatically from the selected Project. Overdue tasks are highlighted in red.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "tasks")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    project_name_by_index = {i + 1: p[1] for i, p in enumerate(PROJECTS)}

    for i, (proj_idx, name, dept, owner, ttype, prio, status, created, due, completed, est, notes) in enumerate(TASKS):
        r = r0 + i
        id_col = col_letter(idx["Task ID"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Task ID"])
        # use a fixed anchor column (Created Date is always populated) for the running count
        anchor_col = col_letter(idx["Created Date"])
        id_cell.value = f'="TASK-"&TEXT(SUBTOTAL(103,${anchor_col}${r0}:{anchor_col}{r}),"0000")'

        vals = {
            "Task Name": name, "Project": project_name_by_index[proj_idx], "Department": dept, "Owner": owner,
            "Task Type": ttype, "Priority": prio, "Status": status, "Created Date": created,
            "Due Date": due, "Completed Date": completed, "Estimated Hours": est, "Notes": notes,
        }
        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field in ("Created Date", "Due Date", "Completed Date"):
                cell.number_format = "yyyy-mm-dd"

        proj_ref = f"{col_letter(idx['Project'])}{r}"
        comp_cell = ws.cell(row=r, column=FIRST_COL + idx["Company"],
                             value=f'=IFERROR(INDEX(Tbl_Projects[Company],MATCH({proj_ref},Tbl_Projects[Project Name],0)),"")')

        task_ref = f"{col_letter(idx['Task ID'])}{r}"
        actual_cell = ws.cell(row=r, column=FIRST_COL + idx["Actual Hours"],
                               value=f'=ROUND(SUMIFS(Tbl_TimeLog[Total Hours],Tbl_TimeLog[Task],{task_ref}),1)')
        actual_cell.number_format = "0.0"

        status_ref = f"{col_letter(idx['Status'])}{r}"
        est_ref = f"{col_letter(idx['Estimated Hours'])}{r}"
        actual_ref = f"{col_letter(idx['Actual Hours'])}{r}"
        pct_cell = ws.cell(row=r, column=FIRST_COL + idx["Progress %"],
                            value=f'=IF({status_ref}="Completed",1,IF(OR({status_ref}="Cancelled",{est_ref}=0),0,MIN(0.95,{actual_ref}/{est_ref})))')
        pct_cell.number_format = "0%"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(TASKS) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Project": "=List_ProjectNames",
        "Department": f"={NAME_MAP['Departments']}",
        "Owner": "=List_EmployeeNames",
        "Task Type": f"={NAME_MAP['Task Types']}",
        "Priority": f"={NAME_MAP['Priority Levels']}",
        "Status": f"={NAME_MAP['Statuses']}",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}5000", formula)

    status_col = col_letter(idx["Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{status_col}{r0}:{status_col}5000",
        {"Not Started": styles.GRAY, "In Progress": styles.INFO, "Blocked": styles.DANGER,
         "In Review": styles.ACCENT_PURPLE, "Completed": styles.SUCCESS,
         "Cancelled": styles.GRAY, "On Hold": styles.WARNING},
    )
    prio_col = col_letter(idx["Priority"])
    styles.apply_status_conditional_formatting(
        ws, f"{prio_col}{r0}:{prio_col}5000",
        {"Critical": styles.DANGER, "High": styles.WARNING, "Medium": styles.INFO, "Low": styles.GRAY},
    )
    pct_col = col_letter(idx["Progress %"])
    styles.add_databar(ws, f"{pct_col}{r0}:{pct_col}5000", color=styles.ACCENT_AMBER)

    # Overdue highlight: Due Date in the past AND status not Completed/Cancelled.
    # Column refs are locked with $ so the check doesn't drift as the rule is
    # evaluated across every column in the row - only the row should shift.
    due_col = col_letter(idx["Due Date"])
    from openpyxl.formatting.rule import FormulaRule
    overdue_rule = FormulaRule(
        formula=[f'AND(${due_col}{r0}<TODAY(),${status_col}{r0}<>"Completed",${status_col}{r0}<>"Cancelled",${due_col}{r0}<>"")'],
        fill=styles.hex_fill(styles.DANGER),
        font=styles.header_font(),
    )
    ws.conditional_formatting.add(f"{col_letter(0)}{r0}:{col_letter(len(COLUMNS)-1)}5000", overdue_rule)

    define_dynamic_range(wb, "List_TaskIDs", "Tasks", col_letter(idx["Task ID"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Task ID'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
