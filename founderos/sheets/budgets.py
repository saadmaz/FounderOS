"""Budgets: one row = one company/category/month budget target. Actual
Spent and Variance are calculated live from Expenses - this table IS the
budget-vs-actual report, not a separate copy of one.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..seed_data import COMPANIES, BUDGETS
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Budgets"

COLUMNS = [
    ("Budget ID", 12), ("Company", 18), ("Category", 18), ("Month", 12),
    ("Budgeted Amount", 15), ("Actual Spent", 13), ("Variance", 12),
    ("% Used", 10), ("Notes", 26),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Budgets")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["companies"]
    common.write_title(ws, "📊  Budgets — Budget vs Actual",
                        "One row per company/category/month. Actual Spent and Variance pull live from Expenses - this table is the budget-vs-actual report.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "companies")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    company_name = {f"COMP-{i+1:04d}": c["name"] for i, c in enumerate(COMPANIES)}

    for i, (comp_id, category, month, budgeted) in enumerate(BUDGETS):
        r = r0 + i
        anchor_col = col_letter(idx["Month"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Budget ID"])
        id_cell.value = f'="BUD-"&TEXT(SUBTOTAL(103,${anchor_col}${r0}:{anchor_col}{r}),"0000")'

        ws.cell(row=r, column=FIRST_COL + idx["Company"], value=company_name[comp_id])
        ws.cell(row=r, column=FIRST_COL + idx["Category"], value=category)
        m_cell = ws.cell(row=r, column=FIRST_COL + idx["Month"], value=month)
        m_cell.number_format = "mmm yyyy"
        b_cell = ws.cell(row=r, column=FIRST_COL + idx["Budgeted Amount"], value=budgeted)
        b_cell.number_format = "#,##0.00"

        comp_ref = f"{col_letter(idx['Company'])}{r}"
        cat_ref = f"{col_letter(idx['Category'])}{r}"
        month_ref = f"{col_letter(idx['Month'])}{r}"
        actual_cell = ws.cell(
            row=r, column=FIRST_COL + idx["Actual Spent"],
            value=(f'=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Company],{comp_ref},'
                   f'Tbl_Expenses[Category],{cat_ref},Tbl_Expenses[Date],">="&{month_ref},'
                   f'Tbl_Expenses[Date],"<="&EOMONTH({month_ref},0)),2)'))
        actual_cell.number_format = "#,##0.00"

        budget_ref = f"{col_letter(idx['Budgeted Amount'])}{r}"
        actual_ref = f"{col_letter(idx['Actual Spent'])}{r}"
        var_cell = ws.cell(row=r, column=FIRST_COL + idx["Variance"],
                            value=f'=ROUND({budget_ref}-{actual_ref},2)')
        var_cell.number_format = "#,##0.00"
        pct_cell = ws.cell(row=r, column=FIRST_COL + idx["% Used"],
                            value=f'=IF({budget_ref}=0,0,{actual_ref}/{budget_ref})')
        pct_cell.number_format = "0%"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(BUDGETS) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Company": "=List_CompanyNames",
        "Category": f"={NAME_MAP['Expense Categories']}",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}2000", formula)

    pct_col = col_letter(idx["% Used"])
    styles.add_databar(ws, f"{pct_col}{r0}:{pct_col}2000", color=styles.WARNING)

    # Over-budget highlight: Actual Spent exceeds Budgeted Amount
    from openpyxl.formatting.rule import FormulaRule
    var_col = col_letter(idx["Variance"])
    rule = FormulaRule(
        formula=[f'${var_col}{r0}<0'],
        fill=styles.hex_fill(styles.DANGER),
        font=styles.header_font(),
    )
    ws.conditional_formatting.add(f"{var_col}{r0}:{var_col}2000", rule)

    ws.freeze_panes = f"{col_letter(idx['Month'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
