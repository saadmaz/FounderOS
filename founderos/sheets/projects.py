"""Projects: joins to Companies by name, joins to Tasks/Time Log for
progress and hours. Owner picked from the Employees master list.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import COMPANIES, PROJECTS
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Projects"

COLUMNS = [
    ("Project ID", 12), ("Company", 20), ("Project Name", 26), ("Description", 34),
    ("Priority", 11), ("Department", 15), ("Owner", 18), ("Status", 13),
    ("Start Date", 12), ("End Date", 12), ("Estimated Hours", 13), ("Actual Hours", 12),
    ("Remaining Hours", 13), ("Total Tasks", 11), ("Completed Tasks", 13),
    ("Completion %", 12), ("Notes", 28),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Projects")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["projects"]
    common.write_title(ws, "📁  Projects",
                        "One row per project. Hours and completion % roll up automatically from Tasks and Time Log.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "projects")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    company_name = {f"COMP-{i+1:04d}": c["name"] for i, c in enumerate(COMPANIES)}

    for i, (comp_id, name, desc, prio, dept, owner, status, start, end, est) in enumerate(PROJECTS):
        r = r0 + i
        name_col = col_letter(idx["Project Name"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Project ID"])
        id_cell.value = f'="PROJ-"&TEXT(SUBTOTAL(103,${name_col}${r0}:{name_col}{r}),"0000")'

        vals = {
            "Company": company_name[comp_id], "Project Name": name, "Description": desc,
            "Priority": prio, "Department": dept, "Owner": owner, "Status": status,
            "Start Date": start, "End Date": end, "Estimated Hours": est, "Notes": "",
        }
        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field in ("Start Date", "End Date"):
                cell.number_format = "yyyy-mm-dd"

        proj_ref = f"{name_col}{r}"
        est_ref = f"{col_letter(idx['Estimated Hours'])}{r}"
        actual_cell = ws.cell(row=r, column=FIRST_COL + idx["Actual Hours"],
                               value=f'=ROUND(SUMIFS(Tbl_TimeLog[Total Hours],Tbl_TimeLog[Project],{proj_ref}),1)')
        actual_cell.number_format = "0.0"
        actual_ref = f"{col_letter(idx['Actual Hours'])}{r}"
        rem_cell = ws.cell(row=r, column=FIRST_COL + idx["Remaining Hours"],
                            value=f'=ROUND(MAX(0,{est_ref}-{actual_ref}),1)')
        rem_cell.number_format = "0.0"

        total_cell = ws.cell(row=r, column=FIRST_COL + idx["Total Tasks"],
                              value=f'=COUNTIFS(Tbl_Tasks[Project],{proj_ref})')
        done_cell = ws.cell(row=r, column=FIRST_COL + idx["Completed Tasks"],
                             value=f'=COUNTIFS(Tbl_Tasks[Project],{proj_ref},Tbl_Tasks[Status],"Completed")')
        total_ref = f"{col_letter(idx['Total Tasks'])}{r}"
        done_ref = f"{col_letter(idx['Completed Tasks'])}{r}"
        pct_cell = ws.cell(row=r, column=FIRST_COL + idx["Completion %"],
                            value=f'=IF({total_ref}=0,0,{done_ref}/{total_ref})')
        pct_cell.number_format = "0%"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(PROJECTS) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Company": "=List_CompanyNames",
        "Priority": f"={NAME_MAP['Priority Levels']}",
        "Department": f"={NAME_MAP['Departments']}",
        "Owner": "=List_EmployeeNames",
        "Status": f"={NAME_MAP['Statuses']}",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}2000", formula)

    status_col = col_letter(idx["Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{status_col}{r0}:{status_col}2000",
        {"Not Started": styles.GRAY, "In Progress": styles.INFO, "Blocked": styles.DANGER,
         "In Review": styles.ACCENT_PURPLE, "Completed": styles.SUCCESS,
         "Cancelled": styles.GRAY, "On Hold": styles.WARNING},
    )
    prio_col = col_letter(idx["Priority"])
    styles.apply_status_conditional_formatting(
        ws, f"{prio_col}{r0}:{prio_col}2000",
        {"Critical": styles.DANGER, "High": styles.WARNING, "Medium": styles.INFO, "Low": styles.GRAY},
    )
    pct_col = col_letter(idx["Completion %"])
    styles.add_databar(ws, f"{pct_col}{r0}:{pct_col}2000", color=styles.ACCENT_PURPLE)

    define_dynamic_range(wb, "List_ProjectNames", "Projects", col_letter(idx["Project Name"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Project Name'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
