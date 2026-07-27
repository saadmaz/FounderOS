"""Invoices: what was billed. Amount Collected / Outstanding are derived
live from whichever Revenue rows reference this Invoice ID - never typed.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import COMPANIES, INVOICES
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Invoices"

COLUMNS = [
    ("Invoice ID", 12), ("Company", 18), ("Client", 22), ("Currency", 10),
    ("Amount", 12), ("Issue Date", 12), ("Due Date", 12), ("Status", 12),
    ("Paid Date", 12), ("Amount Collected", 15), ("Outstanding", 12), ("Notes", 26),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Invoices")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["companies"]
    common.write_title(ws, "📄  Invoices",
                        "One row per invoice. Amount Collected and Outstanding are calculated automatically from the Revenue table - link a Revenue row to an Invoice ID to mark it collected.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "companies")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    company_name = {f"COMP-{i+1:04d}": c["name"] for i, c in enumerate(COMPANIES)}

    for i, row in enumerate(INVOICES):
        comp_id, client, currency, amount, issue_date, due_date, status, paid_date, notes = row
        r = r0 + i
        anchor_col = col_letter(idx["Issue Date"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Invoice ID"])
        id_cell.value = f'="INV-"&TEXT(SUBTOTAL(103,${anchor_col}${r0}:{anchor_col}{r}),"0000")'

        vals = {
            "Company": company_name[comp_id], "Client": client, "Currency": currency,
            "Amount": amount, "Issue Date": issue_date, "Due Date": due_date,
            "Status": status, "Paid Date": paid_date, "Notes": notes,
        }
        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field in ("Issue Date", "Due Date", "Paid Date"):
                cell.number_format = "yyyy-mm-dd"
            if field == "Amount":
                cell.number_format = "#,##0.00"

        # Only Revenue rows actually marked Paid count as collected - a
        # linked-but-unpaid Revenue row means the invoice is still outstanding.
        inv_ref = f"{col_letter(idx['Invoice ID'])}{r}"
        collected_cell = ws.cell(
            row=r, column=FIRST_COL + idx["Amount Collected"],
            value=(f'=ROUND(SUMIFS(Tbl_Revenue[Amount],Tbl_Revenue[Invoice ID],{inv_ref},'
                   f'Tbl_Revenue[Payment Status],"Paid"),2)'))
        collected_cell.number_format = "#,##0.00"
        amt_ref = f"{col_letter(idx['Amount'])}{r}"
        coll_ref = f"{col_letter(idx['Amount Collected'])}{r}"
        outstanding_cell = ws.cell(row=r, column=FIRST_COL + idx["Outstanding"],
                                    value=f'=ROUND(MAX(0,{amt_ref}-{coll_ref}),2)')
        outstanding_cell.number_format = "#,##0.00"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(INVOICES) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Company": "=List_CompanyNames",
        "Currency": f"={NAME_MAP['Currencies']}",
        "Status": f"={NAME_MAP['Invoice Status']}",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}5000", formula)

    status_col = col_letter(idx["Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{status_col}{r0}:{status_col}5000",
        {"Draft": styles.GRAY, "Sent": styles.INFO, "Paid": styles.SUCCESS,
         "Overdue": styles.DANGER, "Cancelled": styles.GRAY},
    )

    define_dynamic_range(wb, "List_InvoiceIDs", "Invoices", col_letter(idx["Invoice ID"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Invoice ID'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
