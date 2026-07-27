"""Finance dashboard: P&L by company, expense breakdown by category, and
the money KPIs that matter across all companies at once. Every number
here is a formula against Companies/Expenses/Revenue/Invoices - nothing
is duplicated from those master tables.
"""
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from .. import styles
from ..seed_data import COMPANIES, LOOKUPS
from . import common
from . import companies as companies_mod
from . import lookups as lookups_mod
from .home import _kpi_card, _button

SNAPSHOT_MAX_ROWS = 25
CATEGORY_LIST = LOOKUPS["Expense Categories"]

KPI_CARDS = [
    ("Total Revenue", '=ROUND(SUM(Tbl_Revenue[Amount]),0)', "companies", "#,##0"),
    ("Total Expenses", '=ROUND(SUM(Tbl_Expenses[Total]),0)', "tasks", "#,##0"),
    ("Net Profit", '=ROUND(SUM(Tbl_Revenue[Amount])-SUM(Tbl_Expenses[Total]),0)', "time", "#,##0"),
    ("Outstanding Reimbursements",
     '=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Reimbursable],"Yes",Tbl_Expenses[Reimbursed],"No"),0)',
     "tasks", "#,##0"),
    ("Outstanding Invoices", '=ROUND(SUM(Tbl_Invoices[Outstanding]),0)', "projects", "#,##0"),
    # Revenue-weighted, not a plain average of each company's % - a tiny
    # pre-launch company with a lopsided margin shouldn't swing the headline number.
    ("Blended Profit Margin",
     '=IFERROR((SUM(Tbl_Revenue[Amount])-SUM(Tbl_Expenses[Total]))/SUM(Tbl_Revenue[Amount]),0)',
     "companies", "0%"),
]


def build(wb):
    ws = wb.create_sheet("Finance")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["companies"]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for c in range(2, 16):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws.merge_cells("B2:O4")
    hero = ws["B2"]
    hero.value = "💰 Finance — P&L, Budgets & Cash Flow"
    hero.font = Font(name=styles.FONT_NAME, size=20, bold=True, color=styles.WHITE)
    hero.fill = styles.hex_fill(styles.NAVY)
    hero.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for r in range(2, 5):
        for c in range(2, 16):
            ws.cell(row=r, column=c).fill = styles.hex_fill(styles.NAVY)
    ws.merge_cells("B5:O5")
    sub = ws["B5"]
    sub.value = "Every figure below is a live formula against Expenses, Revenue, and Invoices - nothing is re-typed here."
    sub.font = styles.body_font(color=styles.GRAY, italic=True, size=9)
    ws.row_dimensions[5].height = 16

    nav_row = 7
    ws.cell(row=nav_row, column=2, value="NAVIGATE").font = Font(name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    nav_row += 1
    nav_targets = [
        ("🧾 Expenses", "Expenses", styles.MODULE_COLORS["tasks"]),
        ("💰 Revenue", "Revenue", styles.MODULE_COLORS["companies"]),
        ("📄 Invoices", "Invoices", styles.ACCENT_PURPLE),
        ("📊 Budgets", "Budgets", styles.ACCENT_AMBER),
        ("🏭 Vendors", "Vendors", styles.ACCENT_BLUE),
        ("🏠 Home", "Home", styles.MODULE_COLORS["home"]),
    ]
    col = 2
    for label, target, color in nav_targets:
        _button(ws, nav_row, col, 2, label, target, color)
        col += 3

    kpi_row = nav_row + 2
    ws.cell(row=kpi_row, column=2, value="KEY METRICS").font = Font(name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    kpi_row += 1
    col = 2
    row_cursor = kpi_row
    for i, (label, formula, module_key, numfmt) in enumerate(KPI_CARDS):
        _kpi_card(ws, row_cursor, col, label, formula, styles.MODULE_COLORS[module_key], number_format=numfmt)
        col += 3
        if (i + 1) % 3 == 0:
            col = 2
            row_cursor += 4

    # --- P&L by Company panel -------------------------------------------
    snap_header_row = row_cursor + 1
    ws.cell(row=snap_header_row, column=2, value="P&L BY COMPANY").font = Font(
        name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    snap_header_row += 1

    headers = ["Company", "Revenue", "Expenses", "Profit", "Margin", "Rev / Hour", "Outstanding Reimb."]
    widths = [20, 13, 13, 12, 10, 12, 15]
    for i, h in enumerate(headers):
        cell = ws.cell(row=snap_header_row, column=2 + i, value=h)
        ws.column_dimensions[get_column_letter(2 + i)].width = widths[i]
    styles.style_table_header_row(ws, snap_header_row, 2, 2 + len(headers) - 1, "companies")

    comp_idx = {name: i for i, (name, _) in enumerate(companies_mod.COLUMNS)}
    comp_r0 = 4
    src_cols = ["Company Name", "Total Revenue", "Total Expenses", "Total Profit",
                "Profit Margin", "Revenue / Hour", "Outstanding Reimbursements"]

    for k in range(1, SNAPSHOT_MAX_ROWS + 1):
        r = snap_header_row + k
        comp_row = comp_r0 + k - 1
        for i, field in enumerate(src_cols):
            src_col_letter = companies_mod.col_letter(comp_idx[field])
            cell = ws.cell(row=r, column=2 + i)
            cell.value = (
                f'=IF(ROW()-{snap_header_row}<=COUNTA(Tbl_Companies[Company ID]),'
                f'IFERROR(Companies!{src_col_letter}{comp_row},""),"")'
            )
            cell.font = styles.body_font()
            cell.border = styles.THIN_BORDER
            if field in ("Total Revenue", "Total Expenses", "Total Profit", "Revenue / Hour", "Outstanding Reimbursements"):
                cell.number_format = "#,##0.00"
            if field == "Profit Margin":
                cell.number_format = "0%"

    pnl_last_row = snap_header_row + SNAPSHOT_MAX_ROWS

    # P&L chart: Revenue vs Expenses per company, straight from Companies sheet
    n_companies = len(COMPANIES)
    comp_ws = wb["Companies"]
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Revenue vs Expenses by Company"
    chart.style = 10
    chart.y_axis.title = "Amount"
    chart.height = 8
    chart.width = 18
    rev_col = 2 + comp_idx["Total Revenue"]
    exp_col = 2 + comp_idx["Total Expenses"]
    data = Reference(comp_ws, min_col=rev_col, max_col=exp_col, min_row=3, max_row=3 + n_companies)
    cats = Reference(comp_ws, min_col=2 + comp_idx["Company Name"], min_row=4, max_row=3 + n_companies)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"J{snap_header_row}")

    # --- Expense Breakdown by Category -----------------------------------
    cat_header_row = pnl_last_row + 20
    ws.cell(row=cat_header_row, column=2, value="EXPENSES BY CATEGORY").font = Font(
        name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    cat_header_row += 1
    cat_headers = ["Category", "Total Spent"]
    for i, h in enumerate(cat_headers):
        ws.cell(row=cat_header_row, column=2 + i, value=h)
    styles.style_table_header_row(ws, cat_header_row, 2, 3, "tasks")

    lk_col_letter = get_column_letter(
        lookups_mod.FIRST_COL + lookups_mod.LIST_ORDER.index("Expense Categories") * 2
    )
    for i in range(len(CATEGORY_LIST)):
        r = cat_header_row + 1 + i
        lk_row = lookups_mod.FIRST_DATA_ROW + i
        cat_cell = ws.cell(row=r, column=2, value=f"=Lookups!{lk_col_letter}{lk_row}")
        cat_cell.font = styles.body_font()
        cat_cell.border = styles.THIN_BORDER
        cat_ref = f"B{r}"
        tot_cell = ws.cell(row=r, column=3,
                            value=f'=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Category],{cat_ref}),2)')
        tot_cell.number_format = "#,##0.00"
        tot_cell.font = styles.body_font()
        tot_cell.border = styles.THIN_BORDER
    cat_last_row = cat_header_row + len(CATEGORY_LIST)

    cat_chart = BarChart()
    cat_chart.type = "bar"
    cat_chart.title = "Expenses by Category (All Companies)"
    cat_chart.style = 10
    cat_chart.height = 10
    cat_chart.width = 18
    cat_data = Reference(ws, min_col=3, min_row=cat_header_row, max_row=cat_last_row)
    cat_cats = Reference(ws, min_col=2, min_row=cat_header_row + 1, max_row=cat_last_row)
    cat_chart.add_data(cat_data, titles_from_data=True)
    cat_chart.set_categories(cat_cats)
    cat_chart.legend = None
    ws.add_chart(cat_chart, f"J{cat_header_row}")

    ws.freeze_panes = "B8"
    return ws
