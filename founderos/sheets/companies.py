"""Companies: the master company registry. Every other module joins back
to this table by Company Name. Adding a new company = adding one row here;
every dashboard and KPI updates on its own.
"""
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import COMPANIES
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Companies"

COLUMNS = [
    ("Company ID", 12), ("Company Name", 22), ("Legal Name", 24), ("Industry", 18),
    ("Business Type", 14), ("Website", 26), ("Email", 24), ("Phone", 16),
    ("Country", 12), ("Currency", 10), ("Founder", 18), ("Status", 12),
    ("Date Started", 13), ("Stage", 14), ("Employees", 11), ("Total Projects", 13),
    ("Active Projects", 13), ("Open Tasks", 11), ("Completed Tasks", 14),
    ("Hours Logged", 12), ("Total Revenue", 13), ("Total Expenses", 13),
    ("Total Profit", 12), ("Profit Margin", 12), ("Monthly Revenue", 14),
    ("Monthly Expenses", 15), ("Monthly Profit", 13), ("Revenue / Hour", 13),
    ("Outstanding Reimbursements", 17), ("Notes", 30),
]

FIRST_COL = 2  # column B


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Companies")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["companies"]
    common.write_title(ws, "🏢  Companies — Master Registry",
                        "One row per company. KPI columns (right side) are calculated automatically from Projects, Tasks, and Time Log.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "companies")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    for i, comp in enumerate(COMPANIES):
        r = r0 + i
        vals = {
            "Company Name": comp["name"], "Legal Name": comp["legal"], "Industry": comp["industry"],
            "Business Type": comp["biz_type"], "Website": comp["website"], "Email": comp["email"],
            "Phone": comp["phone"], "Country": comp["country"], "Currency": comp["currency"],
            "Founder": comp["founder"], "Status": comp["status"], "Date Started": comp["started"],
            "Stage": comp["stage"], "Notes": comp["notes"],
        }
        # Company ID: auto-incrementing calculated column
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Company ID"])
        id_cell.value = f'="COMP-"&TEXT(SUBTOTAL(103,${col_letter(idx["Company Name"])}${r0}:{col_letter(idx["Company Name"])}{r}),"0000")'

        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field == "Date Started":
                cell.number_format = "yyyy-mm-dd"

        # KPI calculated columns
        name_ref = f"{col_letter(idx['Company Name'])}{r}"
        ws.cell(row=r, column=FIRST_COL + idx["Employees"],
                value=f'=COUNTIFS(Tbl_Employees[Primary Company],{name_ref})')
        ws.cell(row=r, column=FIRST_COL + idx["Total Projects"],
                value=f'=COUNTIFS(Tbl_Projects[Company],{name_ref})')
        ws.cell(row=r, column=FIRST_COL + idx["Active Projects"],
                value=f'=COUNTIFS(Tbl_Projects[Company],{name_ref},Tbl_Projects[Status],"<>Completed",Tbl_Projects[Status],"<>Cancelled")')
        ws.cell(row=r, column=FIRST_COL + idx["Open Tasks"],
                value=f'=COUNTIFS(Tbl_Tasks[Company],{name_ref},Tbl_Tasks[Status],"<>Completed",Tbl_Tasks[Status],"<>Cancelled")')
        ws.cell(row=r, column=FIRST_COL + idx["Completed Tasks"],
                value=f'=COUNTIFS(Tbl_Tasks[Company],{name_ref},Tbl_Tasks[Status],"Completed")')
        hrs = ws.cell(row=r, column=FIRST_COL + idx["Hours Logged"],
                       value=f'=ROUND(SUMIFS(Tbl_TimeLog[Total Hours],Tbl_TimeLog[Company],{name_ref}),1)')
        hrs.number_format = "0.0"

        # Finance KPIs - all derived from Tbl_Revenue / Tbl_Expenses, nothing typed.
        rev_cell = ws.cell(row=r, column=FIRST_COL + idx["Total Revenue"],
                            value=f'=ROUND(SUMIFS(Tbl_Revenue[Amount],Tbl_Revenue[Company],{name_ref}),2)')
        rev_cell.number_format = "#,##0.00"
        exp_cell = ws.cell(row=r, column=FIRST_COL + idx["Total Expenses"],
                            value=f'=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Company],{name_ref}),2)')
        exp_cell.number_format = "#,##0.00"
        rev_ref = f"{col_letter(idx['Total Revenue'])}{r}"
        exp_ref = f"{col_letter(idx['Total Expenses'])}{r}"
        profit_cell = ws.cell(row=r, column=FIRST_COL + idx["Total Profit"],
                               value=f'=ROUND({rev_ref}-{exp_ref},2)')
        profit_cell.number_format = "#,##0.00"
        profit_ref = f"{col_letter(idx['Total Profit'])}{r}"
        margin_cell = ws.cell(row=r, column=FIRST_COL + idx["Profit Margin"],
                               value=f'=IF({rev_ref}=0,0,{profit_ref}/{rev_ref})')
        margin_cell.number_format = "0%"

        mrev_cell = ws.cell(
            row=r, column=FIRST_COL + idx["Monthly Revenue"],
            value=(f'=ROUND(SUMIFS(Tbl_Revenue[Amount],Tbl_Revenue[Company],{name_ref},'
                   f'Tbl_Revenue[Date],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1)),2)'))
        mrev_cell.number_format = "#,##0.00"
        mexp_cell = ws.cell(
            row=r, column=FIRST_COL + idx["Monthly Expenses"],
            value=(f'=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Company],{name_ref},'
                   f'Tbl_Expenses[Date],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1)),2)'))
        mexp_cell.number_format = "#,##0.00"
        mrev_ref = f"{col_letter(idx['Monthly Revenue'])}{r}"
        mexp_ref = f"{col_letter(idx['Monthly Expenses'])}{r}"
        mprofit_cell = ws.cell(row=r, column=FIRST_COL + idx["Monthly Profit"],
                                value=f'=ROUND({mrev_ref}-{mexp_ref},2)')
        mprofit_cell.number_format = "#,##0.00"

        hrs_ref = f"{col_letter(idx['Hours Logged'])}{r}"
        rph_cell = ws.cell(row=r, column=FIRST_COL + idx["Revenue / Hour"],
                            value=f'=IF({hrs_ref}=0,0,ROUND({rev_ref}/{hrs_ref},2))')
        rph_cell.number_format = "#,##0.00"

        outstanding_cell = ws.cell(
            row=r, column=FIRST_COL + idx["Outstanding Reimbursements"],
            value=(f'=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Company],{name_ref},'
                   f'Tbl_Expenses[Reimbursable],"Yes",Tbl_Expenses[Reimbursed],"No"),2)'))
        outstanding_cell.number_format = "#,##0.00"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(COMPANIES) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    # Dropdown validations
    styles.add_data_validation_list(ws, f"{col_letter(idx['Currency'])}{r0}:{col_letter(idx['Currency'])}500",
                                     f"={NAME_MAP['Currencies']}")
    styles.add_data_validation_list(ws, f"{col_letter(idx['Status'])}{r0}:{col_letter(idx['Status'])}500",
                                     f"={NAME_MAP['Company Status']}")
    styles.add_data_validation_list(ws, f"{col_letter(idx['Stage'])}{r0}:{col_letter(idx['Stage'])}500",
                                     f"={NAME_MAP['Company Stage']}")

    # Conditional formatting for status
    status_col = col_letter(idx["Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{status_col}{r0}:{status_col}500",
        {"Active": styles.SUCCESS, "Paused": styles.WARNING, "Archived": styles.GRAY, "Exploring": styles.INFO},
    )

    define_dynamic_range(wb, "List_CompanyNames", "Companies", col_letter(idx["Company Name"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Company Name'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
