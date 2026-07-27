"""Vendors: master supplier list. Feeds the Vendor dropdown on Expenses
and rolls up spend history per vendor automatically.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import VENDORS
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Vendors"

COLUMNS = [
    ("Vendor ID", 12), ("Vendor Name", 26), ("Category", 20), ("Contact Name", 16),
    ("Email", 24), ("Phone", 16), ("Payment Terms", 14), ("Expense Count", 12),
    ("Total Spend", 13), ("Notes", 28),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Vendors")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["companies"]
    common.write_title(ws, "🏭  Vendors — Supplier Directory",
                        "One row per vendor. Expense Count and Total Spend are calculated automatically from the Expenses table.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "companies")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW

    for i, (name, category, contact, email, phone, terms, notes) in enumerate(VENDORS):
        r = r0 + i
        name_col = col_letter(idx["Vendor Name"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Vendor ID"])
        id_cell.value = f'="VEN-"&TEXT(SUBTOTAL(103,${name_col}${r0}:{name_col}{r}),"0000")'

        vals = {
            "Vendor Name": name, "Category": category, "Contact Name": contact,
            "Email": email, "Phone": phone, "Payment Terms": terms, "Notes": notes,
        }
        for field, value in vals.items():
            ws.cell(row=r, column=FIRST_COL + idx[field], value=value)

        vend_ref = f"{name_col}{r}"
        ws.cell(row=r, column=FIRST_COL + idx["Expense Count"],
                value=f'=COUNTIFS(Tbl_Expenses[Vendor],{vend_ref})')
        spend_cell = ws.cell(row=r, column=FIRST_COL + idx["Total Spend"],
                              value=f'=ROUND(SUMIFS(Tbl_Expenses[Total],Tbl_Expenses[Vendor],{vend_ref}),2)')
        spend_cell.number_format = "#,##0.00"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(VENDORS) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    styles.add_data_validation_list(ws, f"{col_letter(idx['Category'])}{r0}:{col_letter(idx['Category'])}1000",
                                     f"={NAME_MAP['Expense Categories']}")

    define_dynamic_range(wb, "List_VendorNames", "Vendors", col_letter(idx["Vendor Name"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Vendor Name'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
