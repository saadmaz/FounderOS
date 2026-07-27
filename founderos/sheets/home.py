"""Home: the Founder Dashboard. Pure navigation + read-only rollups —
every number here is a formula pointing at a master table, nothing is
typed or duplicated on this sheet.
"""
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from .. import styles
from ..seed_data import COMPANIES
from . import companies as companies_mod

NAV_ROW_1 = [
    ("🏢 Companies", "Companies", styles.MODULE_COLORS["companies"]),
    ("📁 Projects", "Projects", styles.MODULE_COLORS["projects"]),
    ("✅ Tasks", "Tasks", styles.MODULE_COLORS["tasks"]),
    ("⏱️ Time Log", "Time_Log", styles.MODULE_COLORS["time"]),
    ("🧑‍💼 Employees", "Employees", styles.MODULE_COLORS["employees"]),
]
NAV_ROW_2 = [
    ("💰 Finance", "Finance", styles.ACCENT_TEAL),
    ("🧾 Expenses", "Expenses", styles.MODULE_COLORS["tasks"]),
    ("📄 Invoices", "Invoices", styles.ACCENT_PURPLE),
    ("📊 Budgets", "Budgets", styles.ACCENT_AMBER),
    ("⚙️ Lookups", "Lookups", styles.MODULE_COLORS["lookups"]),
]
# kept for any external import expecting the old flat name
NAV_TARGETS = NAV_ROW_1 + NAV_ROW_2

KPI_CARDS = [
    ("Total Companies", '=COUNTA(Tbl_Companies[Company ID])', "companies", "0"),
    ("Active Companies", '=COUNTIF(Tbl_Companies[Status],"Active")', "companies", "0"),
    ("Total Projects", '=COUNTA(Tbl_Projects[Project ID])', "projects", "0"),
    ("Active Projects", '=COUNTIFS(Tbl_Projects[Status],"<>Completed",Tbl_Projects[Status],"<>Cancelled")', "projects", "0"),
    ("Open Tasks", '=COUNTIFS(Tbl_Tasks[Status],"<>Completed",Tbl_Tasks[Status],"<>Cancelled")', "tasks", "0"),
    ("Overdue Tasks", '=COUNTIFS(Tbl_Tasks[Due Date],"<"&TODAY(),Tbl_Tasks[Status],"<>Completed",Tbl_Tasks[Status],"<>Cancelled")', "tasks", "0"),
    ("Due This Week", '=COUNTIFS(Tbl_Tasks[Due Date],">="&TODAY(),Tbl_Tasks[Due Date],"<="&(TODAY()+7),Tbl_Tasks[Status],"<>Completed",Tbl_Tasks[Status],"<>Cancelled")', "tasks", "0"),
    ("Hours Logged (This Month)", '=ROUND(SUMIFS(Tbl_TimeLog[Total Hours],Tbl_TimeLog[Date],">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1)),1)', "time", "0"),
    ("Total Hours Logged", '=ROUND(SUM(Tbl_TimeLog[Total Hours]),1)', "time", "0"),
]

FINANCE_KPI_CARDS = [
    ("Total Revenue", '=ROUND(SUM(Tbl_Revenue[Amount]),0)', "companies", "#,##0"),
    ("Total Expenses", '=ROUND(SUM(Tbl_Expenses[Total]),0)', "tasks", "#,##0"),
    ("Net Profit", '=ROUND(SUM(Tbl_Revenue[Amount])-SUM(Tbl_Expenses[Total]),0)', "time", "#,##0"),
]

SNAPSHOT_MAX_ROWS = 25


def _button(ws, row, col_start, col_span, label, target_sheet, color):
    c0 = col_start
    c1 = col_start + col_span - 1
    ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
    cell = ws.cell(row=row, column=c0)
    cell.value = f'=HYPERLINK("#\'{target_sheet}\'!B4","{label}")'
    cell.font = Font(name=styles.FONT_NAME, size=11, bold=True, color=styles.WHITE)
    cell.fill = styles.hex_fill(color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def _kpi_card(ws, row, col_start, label, formula, color, number_format="0"):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_start + 2)
    label_cell = ws.cell(row=row, column=col_start, value=label)
    label_cell.font = Font(name=styles.FONT_NAME, size=9, bold=True, color=styles.WHITE)
    label_cell.fill = styles.hex_fill(color)
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=row + 1, start_column=col_start, end_row=row + 2, end_column=col_start + 2)
    val_cell = ws.cell(row=row + 1, column=col_start, value=formula)
    val_cell.number_format = number_format
    val_cell.font = Font(name=styles.FONT_NAME, size=22, bold=True, color=styles.NAVY)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in (row, row + 1, row + 2):
        for c in range(col_start, col_start + 3):
            ws.cell(row=r, column=c).border = styles.THIN_BORDER
    ws.row_dimensions[row].height = 18


def build(wb):
    ws = wb.create_sheet("Home", 0)
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["home"]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    for col, w in {get_column_letter(c): 13 for c in range(2, 16)}.items():
        ws.column_dimensions[col].width = w

    # Hero banner
    ws.merge_cells("B2:O4")
    hero = ws["B2"]
    hero.value = "🚀 FounderOS — Founder Dashboard"
    hero.font = Font(name=styles.FONT_NAME, size=22, bold=True, color=styles.WHITE)
    hero.fill = styles.hex_fill(styles.NAVY)
    hero.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    for r in range(2, 5):
        for c in range(2, 16):
            ws.cell(row=r, column=c).fill = styles.hex_fill(styles.NAVY)
    ws.merge_cells("B5:O5")
    sub = ws["B5"]
    sub.value = '=CONCATENATE("Live snapshot as of ",TEXT(TODAY(),"dddd, mmmm d, yyyy")," — every number below is a live formula, nothing here is typed by hand.")'
    sub.font = styles.body_font(color=styles.GRAY, italic=True, size=9)
    ws.row_dimensions[5].height = 16

    # Navigation - two rows: Operations, then Finance/System
    nav_row = 7
    ws.cell(row=nav_row, column=2, value="NAVIGATE").font = Font(name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    nav_row += 1
    span = 2
    for nav_group in (NAV_ROW_1, NAV_ROW_2):
        col = 2
        for label, target, color in nav_group:
            _button(ws, nav_row, col, span, label, target, color)
            col += span + 1
        nav_row += 2

    # KPI cards
    kpi_row = nav_row + 1
    ws.cell(row=kpi_row, column=2, value="KEY METRICS").font = Font(name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    kpi_row += 1
    col = 2
    per_row = 3
    row_cursor = kpi_row
    for i, (label, formula, module_key, numfmt) in enumerate(KPI_CARDS):
        _kpi_card(ws, row_cursor, col, label, formula, styles.MODULE_COLORS[module_key], number_format=numfmt)
        col += 3
        if (i + 1) % per_row == 0:
            col = 2
            row_cursor += 4

    # Finance snapshot cards
    fin_row = row_cursor + 1
    ws.cell(row=fin_row, column=2, value="FINANCE AT A GLANCE").font = Font(name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    fin_row += 1
    col = 2
    for label, formula, module_key, numfmt in FINANCE_KPI_CARDS:
        _kpi_card(ws, fin_row, col, label, formula, styles.MODULE_COLORS[module_key], number_format=numfmt)
        col += 3
    row_cursor = fin_row + 4

    # Company snapshot panel
    snap_header_row = row_cursor + 1
    ws.cell(row=snap_header_row, column=2, value="COMPANY SNAPSHOT").font = Font(
        name=styles.FONT_NAME, size=9, bold=True, color=styles.GRAY)
    snap_header_row += 1

    headers = ["Company", "Status", "Stage", "Employees", "Projects", "Open Tasks",
               "Hours Logged", "Revenue", "Expenses", "Profit"]
    widths = [22, 12, 14, 11, 10, 11, 13, 13, 13, 12]
    for i, h in enumerate(headers):
        cell = ws.cell(row=snap_header_row, column=2 + i, value=h)
        ws.column_dimensions[get_column_letter(2 + i)].width = widths[i]
    styles.style_table_header_row(ws, snap_header_row, 2, 2 + len(headers) - 1, "home")

    comp_idx = {name: i for i, (name, _) in enumerate(companies_mod.COLUMNS)}
    comp_r0 = 4  # Companies!FIRST_DATA_ROW
    src_cols = ["Company Name", "Status", "Stage", "Employees", "Total Projects", "Open Tasks",
                "Hours Logged", "Total Revenue", "Total Expenses", "Total Profit"]

    for k in range(1, SNAPSHOT_MAX_ROWS + 1):
        r = snap_header_row + k
        comp_row = comp_r0 + k - 1
        for i, field in enumerate(src_cols):
            src_col_letter = companies_mod.col_letter(comp_idx[field])
            cell = ws.cell(row=r, column=2 + i)
            cell.value = (
                f'=IF(ROW()-{snap_header_row}<=COUNTA(Tbl_Companies[Company ID]),'
                f'IFERROR(Companies!{src_col_letter}{comp_row},""),"")'
            )
            cell.font = styles.body_font()
            cell.border = styles.THIN_BORDER
            if field == "Hours Logged":
                cell.number_format = "0.0"
            if field in ("Total Revenue", "Total Expenses", "Total Profit"):
                cell.number_format = "#,##0.00"

    last_snap_row = snap_header_row + SNAPSHOT_MAX_ROWS
    ws.freeze_panes = "B8"

    # Live chart: Hours Logged by Company, sourced straight from Companies sheet
    chart = BarChart()
    chart.type = "col"
    chart.title = "Hours Logged by Company"
    chart.style = 10
    chart.y_axis.title = "Hours"
    chart.x_axis.title = None
    chart.height = 8
    chart.width = 18
    n_companies = len(COMPANIES)
    comp_ws = wb["Companies"]
    data = Reference(comp_ws, min_col=2 + comp_idx["Hours Logged"], min_row=3,
                      max_row=3 + n_companies)
    cats = Reference(comp_ws, min_col=2 + comp_idx["Company Name"], min_row=4,
                      max_row=3 + n_companies)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"M{snap_header_row}")

    return ws
