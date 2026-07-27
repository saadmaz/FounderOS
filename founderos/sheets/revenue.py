"""Revenue: recognized income. Optionally links back to an Invoice ID;
Invoices roll up "Amount Collected" from whatever Revenue rows reference them.
"""
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import COMPANIES, PROJECTS, REVENUE, INVOICES
from . import common
from .lookups import NAME_MAP

TABLE_NAME = "Tbl_Revenue"

COLUMNS = [
    ("Revenue ID", 12), ("Company", 18), ("Project", 22), ("Client / Source", 22),
    ("Category", 18), ("Date", 12), ("Amount", 12), ("Currency", 10),
    ("Payment Status", 13), ("Invoice ID", 12), ("Notes", 28),
]
FIRST_COL = 2


def col_letter(i):
    return get_column_letter(FIRST_COL + i)


def build(wb):
    ws = wb.create_sheet("Revenue")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["companies"]
    common.write_title(ws, "💰  Revenue",
                        "One row per recognized revenue transaction. Link to an Invoice ID when one exists so invoices show what's actually been collected.")

    hr = common.HEADER_ROW
    for i, (name, width) in enumerate(COLUMNS):
        c = FIRST_COL + i
        ws.cell(row=hr, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = width
    styles.style_table_header_row(ws, hr, FIRST_COL, FIRST_COL + len(COLUMNS) - 1, "companies")

    idx = {name: i for i, (name, _) in enumerate(COLUMNS)}
    r0 = common.FIRST_DATA_ROW
    company_name = {f"COMP-{i+1:04d}": c["name"] for i, c in enumerate(COMPANIES)}
    project_name_by_index = {i + 1: p[1] for i, p in enumerate(PROJECTS)}
    invoice_id_by_index = {i + 1: f"INV-{i + 1:04d}" for i in range(len(INVOICES))}

    for i, row in enumerate(REVENUE):
        comp_id, proj_idx, client, category, rdate, amount, currency, pay_status, inv_idx, notes = row
        r = r0 + i
        anchor_col = col_letter(idx["Date"])
        id_cell = ws.cell(row=r, column=FIRST_COL + idx["Revenue ID"])
        id_cell.value = f'="REV-"&TEXT(SUBTOTAL(103,${anchor_col}${r0}:{anchor_col}{r}),"0000")'

        vals = {
            "Company": company_name[comp_id],
            "Project": project_name_by_index.get(proj_idx, ""),
            "Client / Source": client, "Category": category, "Date": rdate, "Amount": amount,
            "Currency": currency, "Payment Status": pay_status,
            "Invoice ID": invoice_id_by_index.get(inv_idx, ""), "Notes": notes,
        }
        for field, value in vals.items():
            cell = ws.cell(row=r, column=FIRST_COL + idx[field])
            cell.value = value
            if field == "Date":
                cell.number_format = "yyyy-mm-dd"
            if field == "Amount":
                cell.number_format = "#,##0.00"

        for c in range(FIRST_COL, FIRST_COL + len(COLUMNS)):
            ws.cell(row=r, column=c).font = styles.body_font()
            ws.cell(row=r, column=c).border = styles.THIN_BORDER

    last_row = r0 + len(REVENUE) - 1
    ref = f"{col_letter(0)}{hr}:{col_letter(len(COLUMNS) - 1)}{last_row}"
    styles.add_table(ws, ref, TABLE_NAME, style="TableStyleMedium9")

    dv_ranges = {
        "Company": "=List_CompanyNames",
        "Project": "=List_ProjectNames",
        "Category": f"={NAME_MAP['Revenue Categories']}",
        "Currency": f"={NAME_MAP['Currencies']}",
        "Payment Status": f"={NAME_MAP['Payment Status']}",
        "Invoice ID": "=List_InvoiceIDs",
    }
    for field, formula in dv_ranges.items():
        col = col_letter(idx[field])
        styles.add_data_validation_list(ws, f"{col}{r0}:{col}20000", formula)

    pay_col = col_letter(idx["Payment Status"])
    styles.apply_status_conditional_formatting(
        ws, f"{pay_col}{r0}:{pay_col}20000",
        {"Unpaid": styles.DANGER, "Paid": styles.SUCCESS, "Partially Paid": styles.WARNING, "Overdue": styles.DANGER},
    )

    define_dynamic_range(wb, "List_RevenueIDs", "Revenue", col_letter(idx["Revenue ID"]), header_row=hr)

    ws.freeze_panes = f"{col_letter(idx['Revenue ID'])}{r0}"
    ws.row_dimensions[hr].height = 32
    return ws
