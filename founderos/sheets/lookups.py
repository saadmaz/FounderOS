"""System > Lookups sheet: every small reference list that drives dropdowns
elsewhere in the workbook. Add a new value here and it is instantly
available everywhere that list is used (via the dynamic named ranges).
"""
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .. import styles
from ..names import define_dynamic_range
from ..seed_data import LOOKUPS

# order controls left-to-right layout; each list gets one data column
# with one blank spacer column to its left (except the first).
LIST_ORDER = [
    "Departments", "Priority Levels", "Statuses", "Company Status",
    "Company Stage", "Task Types", "Currencies", "Time Categories", "Billable",
    "Employee Status", "Expense Categories", "Revenue Categories", "Payment Methods",
    "Approval Status", "Payment Status", "Invoice Status",
]

HEADER_ROW = 3
FIRST_DATA_ROW = HEADER_ROW + 1
FIRST_COL = 2  # column B

# defined-name key -> lookup list name, used by other sheet modules
NAME_MAP = {
    "Departments": "List_Departments",
    "Priority Levels": "List_Priority",
    "Statuses": "List_Statuses",
    "Company Status": "List_CompanyStatus",
    "Company Stage": "List_CompanyStage",
    "Task Types": "List_TaskTypes",
    "Currencies": "List_Currencies",
    "Time Categories": "List_TimeCategories",
    "Billable": "List_Billable",
    "Employee Status": "List_EmployeeStatus",
    "Expense Categories": "List_ExpenseCategories",
    "Revenue Categories": "List_RevenueCategories",
    "Payment Methods": "List_PaymentMethods",
    "Approval Status": "List_ApprovalStatus",
    "Payment Status": "List_PaymentStatus",
    "Invoice Status": "List_InvoiceStatus",
}


def build(wb):
    ws = wb.create_sheet("Lookups")
    ws.sheet_properties.tabColor = styles.MODULE_COLORS["lookups"]

    ws["B1"] = "⚙️  System Lookups — Master Reference Lists"
    ws["B1"].font = Font(name=styles.FONT_NAME, size=14, bold=True, color=styles.NAVY)
    ws.row_dimensions[1].height = 24
    ws["B2"] = "Add a value below to make it available in every dropdown that uses this list. Do not delete the header row."
    ws["B2"].font = styles.body_font(color=styles.GRAY, italic=True, size=9)

    col_widths = {}
    col = FIRST_COL
    for list_name in LIST_ORDER:
        values = LOOKUPS[list_name]
        letter = get_column_letter(col)
        ws.cell(row=HEADER_ROW, column=col, value=list_name)
        styles.style_table_header_row(ws, HEADER_ROW, col, col, "lookups")
        for i, v in enumerate(values):
            cell = ws.cell(row=FIRST_DATA_ROW + i, column=col, value=v)
            cell.font = styles.body_font()
            cell.border = styles.THIN_BORDER
        last_row = FIRST_DATA_ROW + len(values) - 1
        ref = f"{letter}{HEADER_ROW}:{letter}{last_row}"
        styles.add_table(ws, ref, f"Tbl_Lookup_{list_name.replace(' ', '')}", style="TableStyleMedium9")
        define_dynamic_range(wb, NAME_MAP[list_name], "Lookups", letter, header_row=HEADER_ROW)
        col_widths[letter] = max(16, len(list_name) + 2)
        col += 2  # leave a spacer column

    styles.autosize_columns(ws, col_widths)
    ws.column_dimensions["A"].width = 2
    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    return ws
