"""Shared sheet-layout conventions used by every module sheet:
row 1 = big title, row 2 = subtitle/help text, row 3 = table header,
row 4+ = data. Keeping this identical everywhere makes the workbook feel
like one coherent system rather than 10 differently-shaped tabs.
"""
from openpyxl.styles import Font

from .. import styles

TITLE_ROW = 1
SUBTITLE_ROW = 2
HEADER_ROW = 3
FIRST_DATA_ROW = 4


def write_title(ws, title, subtitle, first_col_letter="B"):
    ws[f"{first_col_letter}{TITLE_ROW}"] = title
    ws[f"{first_col_letter}{TITLE_ROW}"].font = Font(
        name=styles.FONT_NAME, size=14, bold=True, color=styles.NAVY
    )
    ws.row_dimensions[TITLE_ROW].height = 24
    ws[f"{first_col_letter}{SUBTITLE_ROW}"] = subtitle
    ws[f"{first_col_letter}{SUBTITLE_ROW}"].font = styles.body_font(color=styles.GRAY, italic=True, size=9)
    ws.column_dimensions["A"].width = 2
    ws.sheet_view.showGridLines = False
